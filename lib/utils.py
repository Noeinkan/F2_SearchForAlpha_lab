# utils module

import os
from datetime import datetime

import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from typing import Dict, Any, List, Tuple
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

import win32com.client
import pythoncom
import time

class TradingStrategyInput:
    def __init__(self, all_tickers_df: pd.DataFrame, default_ticker: str, 
                 default_start_date: str, default_end_date: str, 
                 default_initial_capital: float, signals_df: pd.DataFrame):
        self.all_tickers_df = all_tickers_df
        self.default_ticker = default_ticker
        self.default_start_date = default_start_date
        self.default_end_date = default_end_date
        self.default_initial_capital = default_initial_capital
        self.signals_df = signals_df
        self.user_inputs = {}
        self.root = None
        self.cancelled = False  # New attribute to track if the dialog was cancelled

    def get_user_input(self) -> Dict[str, Any]:
        self.root = tk.Tk()
        self.root.title("Trading Strategy Input")
        self.root.attributes('-topmost', True)

        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self._create_input_fields(main_frame)
        self._create_radio_groups(main_frame)
        self._create_signal_selection(main_frame)

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Submit", command=self._collect_inputs).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self._cancel).pack(side=tk.LEFT, padx=5)  # New Cancel button

        self.root.bind('<Return>', lambda event: self._collect_inputs())
        self.root.protocol("WM_DELETE_WINDOW", self._cancel)  # Handle window close button

        self._center_window()
        self.root.mainloop()

        # Destroy the window after mainloop ends
        if self.root:
            self.root.destroy()

        if self.cancelled:
            return None  # Return None if cancelled
        return tuple(self.user_inputs.values())

    def _create_input_fields(self, parent):
        sorted_tickers = sorted(self.all_tickers_df['Symbol'].tolist())

        self.symbol_var = tk.StringVar(value=self.default_ticker)
        self._create_labeled_widget(parent, "Select Stock:", ttk.Combobox,
                                    textvariable=self.symbol_var, values=sorted_tickers)

        self.start_date_entry = self._create_labeled_widget(parent, "Start Date:", DateEntry,
                                                            width=12, background='darkblue', foreground='white',
                                                            borderwidth=2, date_pattern='yyyy-mm-dd')
        self.start_date_entry.set_date(self.default_start_date)

        self.end_date_entry = self._create_labeled_widget(parent, "End Date:", DateEntry,
                                                          width=12, background='darkblue', foreground='white',
                                                          borderwidth=2, date_pattern='yyyy-mm-dd')
        self.end_date_entry.set_date(self.default_end_date)

        self.initial_capital_entry = self._create_labeled_widget(parent, "Initial Capital:", ttk.Entry)
        self.initial_capital_entry.insert(0, str(self.default_initial_capital))

    def _create_radio_groups(self, parent):
        self.export_var = tk.StringVar(value="no")
        self.plot_var = tk.StringVar(value="no")
        self.buy_signals_var = tk.StringVar(value="yes")
        self.sell_signals_var = tk.StringVar(value="yes")

        self._create_radio_group(parent, "Export results to .xlsx?", self.export_var)
        self._create_radio_group(parent, "Plot chart in .jpg?", self.plot_var)

        self.conditional_frame = ttk.Frame(parent)
        self._create_radio_group(self.conditional_frame, "Show buy signals?", self.buy_signals_var)
        self._create_radio_group(self.conditional_frame, "Show sell signals?", self.sell_signals_var)

        self.plot_var.trace_add("write", self._toggle_signal_options)
        self._toggle_signal_options()

    def _create_signal_selection(self, parent):
        all_signals = self._get_signals_from_df()

        signals_frame = ttk.LabelFrame(parent, text="Select Signals")
        signals_frame.pack(fill=tk.X, pady=5)

        buy_frame = ttk.LabelFrame(signals_frame, text="Buy Signals")
        buy_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        sell_frame = ttk.LabelFrame(signals_frame, text="Sell Signals")
        sell_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buy_vars = {}
        self.sell_vars = {}

        for signal in all_signals['buy']:
            self._create_signal_radio(buy_frame, signal, self.buy_vars)
        
        for signal in all_signals['sell']:
            self._create_signal_radio(sell_frame, signal, self.sell_vars)

    def _get_signals_from_df(self) -> Dict[str, List[str]]:
        buy_signals = [signal for signal in self.signals_df if 'buy' in signal.lower()]
        sell_signals = [signal for signal in self.signals_df if 'sell' in signal.lower()]
        return {'buy': buy_signals, 'sell': sell_signals}

    def _create_labeled_widget(self, parent, label_text, widget_class, **kwargs):
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=5)
        ttk.Label(frame, text=label_text).pack(side=tk.LEFT)
        widget = widget_class(frame, **kwargs)
        widget.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        return widget

    def _create_radio_group(self, parent, label_text, variable):
        frame = ttk.LabelFrame(parent, text=label_text)
        frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(frame, text="Yes", variable=variable, value="yes").pack(side=tk.LEFT)
        ttk.Radiobutton(frame, text="No", variable=variable, value="no").pack(side=tk.LEFT)

    def _create_signal_radio(self, parent, signal, var_dict):
        var = tk.StringVar(value="no")
        var_dict[signal] = var
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=2)
        ttk.Label(frame, text=signal).pack(side=tk.LEFT)
        ttk.Radiobutton(frame, text="Yes", variable=var, value="yes").pack(side=tk.RIGHT)
        ttk.Radiobutton(frame, text="No", variable=var, value="no").pack(side=tk.RIGHT)

    def _toggle_signal_options(self, *args):
        if self.plot_var.get() == "yes":
            self.conditional_frame.pack(fill=tk.X, pady=5)
        else:
            self.conditional_frame.pack_forget()

    def _collect_inputs(self):
        try:
            initial_capital = float(self.initial_capital_entry.get())
        except ValueError:
            messagebox.showerror("Error", "Initial capital must be a valid number.")
            return

        self.user_inputs = {
            'symbol': self.symbol_var.get(),
            'start_date': self.start_date_entry.get_date().strftime('%Y-%m-%d'),
            'end_date': self.end_date_entry.get_date().strftime('%Y-%m-%d'),
            'initial_capital': initial_capital,
            'export_to_xlsx': self.export_var.get() == "yes",
            'plot_chart': self.plot_var.get() == "yes",
            'show_buy_signals': self.buy_signals_var.get() == "yes",
            'show_sell_signals': self.sell_signals_var.get() == "yes",
            'buy_indicators': [signal for signal, var in self.buy_vars.items() if var.get() == "yes"],
            'sell_indicators': [signal for signal, var in self.sell_vars.items() if var.get() == "yes"]
        }
        self.root.quit()  # This ends the mainloop

    def _cancel(self):
        self.cancelled = True
        self.root.quit()  # End the mainloop

    def _center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry('{}x{}+{}+{}'.format(width, height, x, y))

def get_user_input(all_tickers_df: pd.DataFrame, default_ticker: str, 
                   default_start_date: str, default_end_date: str, 
                   default_initial_capital: float, signals_list: List[str]) -> Tuple[str, str, str, float, bool, bool, bool, bool, List[str], List[str]]:
    input_dialog = TradingStrategyInput(all_tickers_df, default_ticker, default_start_date, 
                                        default_end_date, default_initial_capital, signals_list)
    
    
    result = input_dialog.get_user_input()
    if result is None:
        return None  
    return result


def export_priceaction_to_excel(ticker: str, results: pd.DataFrame, export_type: str, output_dir: str = None) -> str:
    #timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{ticker}_{export_type}_analysis_results.xlsx"
    
    if output_dir:
        file_name = os.path.join(output_dir, file_name)
    
    # Check if the file already exists
    if os.path.exists(file_name):
        # Try to close the file if it's open
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            for workbook in excel.Workbooks:
                if workbook.FullName == os.path.abspath(file_name):
                    workbook.Close()
            excel.Quit()
        except Exception as e:
            print(f"Error closing Excel file: {e}")
        finally:
            pythoncom.CoUninitialize()

        # Load existing workbook to keep formatting
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        
        # Clear existing data
        ws.delete_rows(1, ws.max_row)
    else:
        # Create new workbook if file doesn't exist
        wb = openpyxl.Workbook()
        ws = wb.active

    # Write new data
    for r in dataframe_to_rows(results, index=True, header=True):
        ws.append(r)
    
    # Apply styling (you can customize this part based on your needs)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Save the workbook
    wb.save(file_name)

    # Open the file
    try:
        os.startfile(file_name)
    except Exception as e:
        print(f"Error opening Excel file: {e}")

    return file_name